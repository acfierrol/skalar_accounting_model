"""Excel writer: values-only output that round-trips and carries #NUM! for non-converging XIRR."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import golden_data as gold  # type: ignore[import-not-found]  # test-local fixture module
import pytest
from openpyxl import load_workbook

from skalar_accounting import (
    AccountingSummary,
    BookKind,
    BookReport,
    amortize,
    build_summary,
    consolidate,
    write_workbook,
)
from skalar_capital_mechanics import (
    GCDates,
    NettingInstruction,
    TransactedLedgerRow,
    build_downstream_cash_events,
    build_netting,
    build_transacted_ledger,
    derive_debt_taken,
)


def _reports() -> tuple[AccountingSummary, list[BookReport]]:
    jg, tg = amortize(gold.june_debt_given()), amortize(gold.july_debt_given())
    jt, tt = amortize(gold.june_debt_taken()), amortize(gold.july_debt_taken())
    given_cons = consolidate([jg, tg])
    taken_cons = consolidate([jt, tt])
    given = BookReport(
        kind=BookKind.DEBT_GIVEN, title="Debt Given By Skalar Cohort Led",
        rate=Decimal("0.25"), loans=(jg, tg), consolidated=given_cons,
    )
    taken = BookReport(
        kind=BookKind.DEBT_TAKEN, title="Debt Taken by Skalar Cohort Led",
        rate=Decimal("0.16"), loans=(jt, tt), consolidated=taken_cons,
    )
    return build_summary(given_cons, taken_cons), [given, taken]


def _structure() -> tuple[list[TransactedLedgerRow], list[NettingInstruction]]:
    events = build_downstream_cash_events(
        company_id="SK011", cohort_month=date(2026, 6, 1), counterparty="Kindroid",
        funding_amount=Decimal("160000"), funding_date=date(2026, 6, 8),
        sharing=[(date(2026, 7, 28), Decimal("84000"))],
    )
    derived = derive_debt_taken(events, GCDates(cohort_month=date(2026, 6, 1),
                                                funding_date=date(2026, 6, 5)))
    return build_transacted_ledger(events + derived), build_netting(events + derived)


def test_workbook_round_trips_values(tmp_path: Path) -> None:
    summary, books = _reports()
    ledger, netting = _structure()
    out = write_workbook(tmp_path / "model.xlsx", summary=summary, books=books,
                         ledger=ledger, netting=netting)

    wb = load_workbook(out)  # no data_only: values are written, not formulas
    assert wb.sheetnames == [
        "Summary", "Debt Given By Skalar Cohort Led", "Debt Taken by Skalar Cohort Led", "Structure"
    ]

    summary_ws = wb["Summary"]
    # Row 5 = Revenue; column D (4) is period 0, E (5) period 1.
    revenue_e = summary_ws.cell(row=5, column=5).value
    assert revenue_e == pytest.approx(gold.SUMMARY_REVENUE[1], abs=1e-6)
    # Row 3 = Date (Given); first period is the 2026-06-08 disbursement.
    written_date = summary_ws.cell(row=3, column=4).value
    assert isinstance(written_date, datetime) and written_date.date() == date(2026, 6, 8)
    # Cash impact row 9, last period = 0.
    assert summary_ws.cell(row=9, column=10).value == pytest.approx(
        gold.SUMMARY_CASH_IMPACT[6], abs=1e-6
    )


def test_pure_accrual_loan_writes_num_error(tmp_path: Path) -> None:
    summary, books = _reports()
    out = write_workbook(tmp_path / "model.xlsx", summary=summary, books=books)
    given_ws = load_workbook(out)["Debt Given By Skalar Cohort Led"]
    cells = [c.value for row in given_ws.iter_rows() for c in row]
    assert "#NUM!" in cells  # the July loan's XIRR does not converge


def test_outstanding_written_to_book_sheet(tmp_path: Path) -> None:
    summary, books = _reports()
    out = write_workbook(tmp_path / "model.xlsx", summary=summary, books=books)
    taken_ws = load_workbook(out)["Debt Taken by Skalar Cohort Led"]
    # Find the Consolidated block's "Principal Outstanding" row and check the last period.
    target = None
    for row in taken_ws.iter_rows():
        labels = [c.value for c in row]
        if "Principal Outstanding" in labels:
            target = row
    assert target is not None
    values = [c.value for c in target if isinstance(c.value, (int, float))]
    assert values[-1] == pytest.approx(gold.TAKEN_CONS_OUTSTANDING[6], abs=1e-6)
