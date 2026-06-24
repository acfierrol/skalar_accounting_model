"""Values-only Excel writer (foundation §1, KB workbook layout).

Emits Summary + one sheet per book (per-loan blocks then a Consolidated block) + a Structure
sheet (transacted ledger + netting "to-do"). **No Excel formulas** — every number is computed
in Python (openpyxl ``data_only`` round-trips cleanly). ``None`` XIRRs render as ``#NUM!``,
mirroring the workbook.
"""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from skalar_capital_mechanics import NettingInstruction, TransactedLedgerRow

from .models import (
    AccountingSummary,
    AmortizationSchedule,
    BookReport,
    ConsolidatedBook,
)

_FIRST_VALUE_COL = 4  # column D — labels live in B/C, period values start at D
_NUM = float  # values-only: store Decimals as floats (the workbook is float-valued)


def _write_period_row(ws: Worksheet, row: int, label: str, values: Sequence[object]) -> None:
    ws.cell(row=row, column=3, value=label)
    for offset, value in enumerate(values):
        ws.cell(row=row, column=_FIRST_VALUE_COL + offset, value=value)


def _schedule_block(ws: Worksheet, start: int, title: str, sched: AmortizationSchedule) -> int:
    ws.cell(row=start, column=2, value=title)
    rows = sched.rows
    _write_period_row(ws, start + 1, "Date", [r.date for r in rows])
    _write_period_row(ws, start + 2, "Inflows", [_NUM(r.inflow) for r in rows])
    _write_period_row(ws, start + 3, "Outflows", [_NUM(r.outflow) for r in rows])
    _write_period_row(ws, start + 4, "Principal", [_NUM(r.principal) for r in rows])
    _write_period_row(ws, start + 5, "Interest", [_NUM(r.interest) for r in rows])
    _write_period_row(ws, start + 6, "Net", [_NUM(r.net) for r in rows])
    _write_period_row(ws, start + 7, "Principal Outstanding", [_NUM(r.outstanding) for r in rows])
    ws.cell(row=start + 8, column=3, value="IRR")
    ws.cell(row=start + 8, column=4, value=sched.xirr if sched.xirr is not None else "#NUM!")
    return start + 10


def _consolidated_block(ws: Worksheet, start: int, book: ConsolidatedBook) -> int:
    ws.cell(row=start, column=2, value="Consolidated Loans")
    rows = book.rows
    _write_period_row(ws, start + 1, "Date", [r.date for r in rows])
    _write_period_row(ws, start + 2, "Inflows", [_NUM(r.inflow) for r in rows])
    _write_period_row(ws, start + 3, "Outflows", [_NUM(r.outflow) for r in rows])
    _write_period_row(ws, start + 4, "Principal", [_NUM(r.principal) for r in rows])
    _write_period_row(ws, start + 5, "Interest", [_NUM(r.interest) for r in rows])
    _write_period_row(ws, start + 6, "Net", [_NUM(r.net) for r in rows])
    _write_period_row(ws, start + 7, "Principal Outstanding", [_NUM(r.outstanding) for r in rows])
    ws.cell(row=start + 8, column=3, value="IRR")
    ws.cell(row=start + 8, column=4, value=book.xirr if book.xirr is not None else "#NUM!")
    return start + 10


def _write_book_sheet(wb: Workbook, report: BookReport) -> None:
    ws = wb.create_sheet(title=report.title[:31])  # Excel caps sheet titles at 31 chars
    ws.cell(row=2, column=2, value="IRR for EIR Method")
    ws.cell(row=2, column=3, value=_NUM(report.rate))
    row = 4
    for loan in report.loans:
        row = _schedule_block(ws, row, loan.name, loan)
    _consolidated_block(ws, row, report.consolidated)


def _write_summary_sheet(ws: Worksheet, summary: AccountingSummary) -> None:
    cols = summary.columns
    ws.cell(row=2, column=2, value="Summary")
    _write_period_row(ws, 3, "Date (Given)", [c.date_given for c in cols])
    _write_period_row(ws, 4, "Date (Taken)", [c.date_taken for c in cols])
    _write_period_row(ws, 5, "Revenue", [_NUM(c.revenue) for c in cols])
    _write_period_row(ws, 6, "Cost of Capital (COGS)", [_NUM(c.cost_of_capital) for c in cols])
    _write_period_row(
        ws, 7, "Outstanding Principal Lended", [_NUM(c.outstanding_lended) for c in cols]
    )
    _write_period_row(
        ws, 8, "Outstanding Principal Borrowed", [_NUM(c.outstanding_borrowed) for c in cols]
    )
    _write_period_row(ws, 9, "Skalar Period Cash Impact", [_NUM(c.cash_impact) for c in cols])
    _write_period_row(ws, 10, "Check", [_NUM(c.check) for c in cols])


def _write_structure_sheet(
    ws: Worksheet,
    ledger: Sequence[TransactedLedgerRow],
    netting: Sequence[NettingInstruction],
) -> None:
    ws.cell(row=2, column=2, value="Transacted USD Amount (neg from Skalar, pos to Skalar)")
    headers = ["Amount", "Loan Cohort", "Counterparty", "Date", "Type"]
    for offset, header in enumerate(headers):
        ws.cell(row=3, column=2 + offset, value=header)
    row = 4
    for entry in ledger:
        ws.cell(row=row, column=2, value=_NUM(entry.amount))
        ws.cell(row=row, column=3, value=entry.loan_cohort)
        ws.cell(row=row, column=4, value=entry.counterparty)
        ws.cell(row=row, column=5, value=entry.date)
        ws.cell(row=row, column=6, value=entry.type)
        row += 1

    row += 1
    ws.cell(row=row, column=2, value='Payments "ToDo" (netting per counterparty / date)')
    net_headers = ["Counterparty", "Date", "Net Amount", "Direction"]
    for offset, header in enumerate(net_headers):
        ws.cell(row=row + 1, column=2 + offset, value=header)
    row += 2
    for wire in netting:
        ws.cell(row=row, column=2, value=wire.counterparty)
        ws.cell(row=row, column=3, value=wire.date)
        ws.cell(row=row, column=4, value=_NUM(wire.net_amount))
        ws.cell(row=row, column=5, value=wire.direction)
        row += 1


def write_workbook(
    path: str | Path,
    *,
    summary: AccountingSummary,
    books: Sequence[BookReport],
    ledger: Sequence[TransactedLedgerRow] = (),
    netting: Sequence[NettingInstruction] = (),
) -> Path:
    """Write the accounting workbook (values only) and return the path."""
    wb = Workbook()
    summary_ws = wb.active
    if summary_ws is None:  # pragma: no cover - openpyxl always provides one
        summary_ws = wb.create_sheet()
    summary_ws.title = "Summary"
    _write_summary_sheet(summary_ws, summary)
    for report in books:
        _write_book_sheet(wb, report)
    _write_structure_sheet(wb.create_sheet(title="Structure"), ledger, netting)

    out = Path(path)
    wb.save(out)
    return out
