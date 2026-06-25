"""Pipeline orchestration: event->book gridding, cached load, end-to-end workbook + report."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from skalar_accounting import BookKind, DayCount
from skalar_capital_mechanics import CashEvent, CashEventKind
from skalar_cli.pipeline import (
    books_from_cash_events,
    load_engine_output,
    run_pipeline,
)

FIXTURE = Path(__file__).parents[1] / "fixtures" / "sk011.json"
JUN1 = date(2026, 6, 1)
JUN8, JUN28 = date(2026, 6, 8), date(2026, 6, 28)
JUL28, AUG28 = date(2026, 7, 28), date(2026, 8, 28)


def _ev(kind: CashEventKind, when: date, amount: str) -> CashEvent:
    return CashEvent(
        company_id="SK011", cohort_month=JUN1, date=when, amount=Decimal(amount),
        kind=kind, counterparty="Kindroid",
    )


def test_books_from_cash_events_buckets_legs_over_grid() -> None:
    events = [
        _ev(CashEventKind.FUND_DOWN, JUN8, "-160000"),
        _ev(CashEventKind.ADJUST, JUN28, "-12000"),
        _ev(CashEventKind.SHARE_UP, JUL28, "84000"),
    ]
    grid = [JUN8, JUN28, JUL28, AUG28]  # AUG28 is a pure-accrual tail period
    books = books_from_cash_events(
        events, kind=BookKind.DEBT_GIVEN, rate=Decimal("0.25"),
        day_count=DayCount.MONTH_DAYS_365, grid=grid,
    )
    assert len(books) == 1
    flows = {f.date: (f.inflow, f.outflow) for f in books[0].flows}
    assert flows[JUN8] == (Decimal("0"), Decimal("-160000"))  # funding -> outflow leg
    assert flows[JUN28] == (Decimal("0"), Decimal("-12000"))  # adjustment -> outflow leg
    assert flows[JUL28] == (Decimal("84000"), Decimal("0"))  # sharing -> inflow leg
    assert flows[AUG28] == (Decimal("0"), Decimal("0"))  # zero-filled tail
    assert books[0].name == "June 2026 Loan"


def test_load_engine_output_parses_fixture() -> None:
    engine = load_engine_output(FIXTURE)
    assert engine.company_id == "SK011"
    assert len(engine.given_loans) == 2
    assert len(engine.taken_loans) == 2
    assert engine.given_loans[0].rate == Decimal("0.25")
    assert engine.taken_loans[0].rate == Decimal("0.16")
    # First given flow is the -160000 disbursement on 2026-06-08.
    first = engine.given_loans[0].flows[0]
    assert first.date == JUN8
    assert first.outflow == Decimal("-160000")


def test_run_pipeline_reproduces_workbook_and_report(tmp_path: Path) -> None:
    engine = load_engine_output(FIXTURE)
    out = tmp_path / "model.xlsx"
    report = run_pipeline(engine, out)

    # Headline figures match the workbook consolidation/summary.
    assert float(report.revenue) == pytest.approx(35110.221038056828, abs=1e-4)
    assert float(report.cost_of_capital) == pytest.approx(-16604.164836209857, abs=1e-4)
    assert float(report.outstanding_lended) == pytest.approx(-219610.22103805683, abs=1e-4)
    assert float(report.outstanding_borrowed) == pytest.approx(202196.19436460373, abs=1e-4)
    assert float(report.max_check) == pytest.approx(0.0, abs=1e-6)
    assert report.compliance_violations == 0
    assert report.threshold_breaches == 0
    assert report.output_path == str(out)

    # The workbook is written with the four sheets and re-reads cleanly.
    wb = load_workbook(out)
    assert wb.sheetnames == [
        "Summary", "Debt Given By Skalar Cohort Led", "Debt Taken by Skalar Cohort Led", "Structure"
    ]
