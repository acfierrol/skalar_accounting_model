"""skalar-accounting CLI: end-to-end cached run + argument validation (offline, no network)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from typer.testing import CliRunner

from skalar_cli import run_cmd

runner = CliRunner()
FIXTURE = Path(__file__).parents[1] / "fixtures" / "sk011.json"


def test_run_produces_workbook_and_report(tmp_path: Path) -> None:
    out = tmp_path / "sk011.xlsx"
    result = runner.invoke(
        run_cmd.app,
        ["run", "--company", "SK011", "--cache", str(FIXTURE), "--out", str(out)],
    )
    assert result.exit_code == 0, result.output
    assert out.exists()
    assert "Run report — SK011" in result.output
    assert "Revenue" in result.output
    assert "|Check|" in result.output

    wb = load_workbook(out)
    assert "Summary" in wb.sheetnames
    # Revenue row 5, period E (column 5) is the workbook's 2961.57.
    assert wb["Summary"].cell(row=5, column=5).value is not None


def test_run_creates_missing_output_directory(tmp_path: Path) -> None:
    # The documented command writes to build/…; a not-yet-existing nested out dir must be created.
    out = tmp_path / "build" / "nested" / "sk011.xlsx"
    result = runner.invoke(
        run_cmd.app,
        ["run", "--company", "SK011", "--cache", str(FIXTURE), "--out", str(out)],
    )
    assert result.exit_code == 0, result.output
    assert out.exists()


def test_run_requires_cache_until_live_build_is_wired(tmp_path: Path) -> None:
    result = runner.invoke(
        run_cmd.app, ["run", "--company", "SK011", "--out", str(tmp_path / "x.xlsx")]
    )
    assert result.exit_code != 0
    assert "live BigQuery runs are not yet wired" in result.output


def test_run_rejects_company_mismatch(tmp_path: Path) -> None:
    result = runner.invoke(
        run_cmd.app,
        ["run", "--company", "SK999", "--cache", str(FIXTURE), "--out", str(tmp_path / "x.xlsx")],
    )
    assert result.exit_code != 0
    assert "SK011" in result.output
